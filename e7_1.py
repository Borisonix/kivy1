import kivy

kivy.require('1.0.8')
import openpyxl
from kivy.app import App
from kivy.uix.button import Button
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.app import runTouchApp


class AllApp(App):
    layout = BoxLayout(orientation='horizontal')
    layout_questions = GridLayout(cols=1, size_hint=(None, None), width=400, height=5500)

    def get_questions(self, subject):           # вычитываем вопросы из xls-файла на заданную тему
        print(subject)
        wb = openpyxl.load_workbook('data.xlsx')
        sheet = wb[subject]                     # находим лист с названием искомой темы
        data = []
        for row in range(2, sheet.max_row + 1):                 # перебор всех строк на листе
            question = sheet.cell(row=row, column=1).value      # читаем содержимое ячейки в первом столбце
            data.append(question)                               # и добавляем его в список data
        return self.add_questions(data)

    def add_questions(self, data):              # список вопросов data превращаем в кнопки
        self.layout_questions.clear_widgets()   # удаляем все виджеты в макете layout_questions
        for question in data:
            question = Button(text=question, size=(480, 40), size_hint=(1, None))   # создаем кнопку
            self.layout_questions.add_widget(question)                     # добавляем её в макет layout_questions

    def build(self):
        layout_subjects = GridLayout(cols=1, size_hint=(None, None), width=400, height=1000)    # макет под кнопки тем
        subjects = ['Logic Building',
                    'SQL Server',
                    'Advanced Excel',
                    'Programming in Java',
                    'HTML5, CSS, JavaScript, JQuery',
                    'Web Development using Servlet and JSP',
                    'Android Development using Java',
                    'Hibernate, Spring and JSF',
                    'Application Testing using JUnit ',
                    'Programming using C# ',
                    'Web development using .Net Framework',
                    'Cross Platform app for Microsoft PlayStore',
                    'distributed application with .net framwork',
                    'Machine Learning using python',
                    'Big Data using R and Python']
        subjects_d = {}
        for subject in subjects:
            subjects_d[subject] = subject   # наполняем словарь. ключ = значению
        print(subjects_d)
        for k, v in subjects_d.items():     # перебор всех пар ключ (k) - значение (v)
            # создаем кнопку
            k = Button(text=v, size=(480, 40), size_hint=(1, None))
            # связываем событие on_press кнопки с функцией вычитывания вопросов по теме и по цепочке создания кнопок справа
            k.bind(on_press=lambda k: self.get_questions(k.text))
            layout_subjects.add_widget(k)  # довляем кнопку в макет layout_subjects

        root = BoxLayout()                              # создаем корневой виджет
        scroll_subject = ScrollView(size_hint=(1, 1))   # создаем два виджета типа ScrollView
        scroll_question = ScrollView(size_hint=(1, 1))
        root.add_widget(self.layout)                    # добавляем на root виджет layout
        self.layout.add_widget(scroll_subject)          # в виджет layout добавляем скроллируемый список тем scroll_subject
        scroll_subject.add_widget(layout_subjects)      # а уже в него добавляем макет с кнопками тем
        self.layout.add_widget(scroll_question)         # в виджет layout добавляем скроллируемый список вопросов scroll_question
        scroll_question.add_widget(self.layout_questions)   # а уже в него добавляем макет с кнопками вопросов
        # runTouchApp(scroll_subject)
        # runTouchApp(scroll_question)
        return root


if __name__ == '__main__':
    AllApp().run()
